# This is extract_us_code_citations_2025-03-18.py


import requests
import PyPDF2
import re
import os
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def download_pdf(url):
    """Downloads a PDF from the given URL and saves it as a temporary file."""
    try:
        response = requests.get(url, stream=True, timeout=10)
        response.raise_for_status()
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        for chunk in response.iter_content(chunk_size=1024):
            temp_file.write(chunk)
        temp_file.close()
        print(f"Downloaded {url}")
        return temp_file.name
    except Exception as e:
        print(f"Failed to download {url}: {e}")
        return None


def sanitize_text(text):
    """Removes line breaks and carriage returns from a given text."""
    return re.sub(r"[\r\n]+", " ", text).strip()


def clean_citation(citation):
    """Cleans up citation syntax to match standard formats."""
    citation = re.sub(r"\b(\d+)\s*(U\.S\.C\.|USC)\s*(\d+)\b", r"\1 USC \3", citation)
    citation = re.sub(r"\b(\d+)\s*(C\.F\.R\.|CFR)\s*(\d+)\b", r"\1 CFR \3", citation)
    citation = re.sub(r"\b(E\.O\.|Executive\s*Order)\s*(\d+)\b", r"Executive Order \2", citation)
    citation = re.sub(r"\bEO\s+(\d+)\b", r"EO \1", citation)  # Ensure EO includes the number
    return citation


def extract_toc(reader):
    """Attempts to extract the Table of Contents (TOC) from the PDF."""
    toc = []
    toc_pattern = r"(?P<heading>.+?)\s+(\d+)"
    for page_num, page in enumerate(reader.pages[:10]):  # Check the first 10 pages for a TOC
        text = page.extract_text()
        if text and "Table of Contents" in text:
            matches = re.findall(toc_pattern, text)
            for match in matches:
                heading = sanitize_text(match[0])
                page_start = int(match[1])
                toc.append((heading, page_start))
    return toc


def infer_section_name(toc, page_num, context, page_text):
    """Infers the section name based on TOC or contextual headers."""
    if toc:
        for i, (section, start_page) in enumerate(toc):
            if i + 1 < len(toc) and toc[i + 1][1] > page_num >= start_page:
                return section
            elif i == len(toc) - 1 and page_num >= start_page:  # Last TOC entry
                return section


    # Fallback: Find the nearest one-line paragraph before the context
    lines = page_text.splitlines()
    context_start = page_text.find(context)
    for i in range(len(lines) - 1, -1, -1):
        if len(lines[i].strip()) > 0 and lines[i].strip() in page_text[:context_start]:
            return sanitize_text(lines[i])
    return "Unknown Section"


def extract_us_code_citations(pdf_path, url):
    """Extracts U.S. Code, CFR, and Executive Order citations from a given PDF file."""
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            toc = extract_toc(reader)
            num_pages = len(reader.pages)


            citations = []
            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text = page.extract_text()


                if text:
                    # Updated pattern for flexible "Executive Order" detection and explicit "EO " followed by a number
                    citation_pattern = r"\b(\d+)\s*(U\.S\.C\.|USC|U.S. Code)\s*\u00a7?\s*(\d+(\.\d+)*([a-zA-Z0-9]*)?)|" \
                                       r"\b(\d+)\s*(C\.F\.R\.|CFR|Code of Federal Regulations)\s*\u00a7?\s*(\d+(\.\d+)*([a-zA-Z0-9]*)?)|" \
                                       r"(E\.O\.|Executive\s*Order)\s*(\d+)|" \
                                       r"\bEO\s+(\d+)\b"


                    matches = re.finditer(citation_pattern, text, re.IGNORECASE)


                    for match in matches:
                        citation_text = match.group(0)
                        citation_number = match.group(10) or match.group(8)  # Capture EO number
                        if citation_number:
                            citation_text = f"EO {citation_number}"


                        citation = clean_citation(citation_text)
                        start, end = match.start(), match.end()
                        context = sanitize_text(text[max(0, start - 100):min(len(text), end + 100)])
                        section_name = infer_section_name(toc, page_num + 1, context, text)
                        citation_page_url = f"{url}#page={page_num + 1}"
                        citations.append((citation, citation_page_url, section_name, context, url))


            return citations
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
        return []


def process_url(url):
    """Handles downloading, extracting, and cleaning up for a single URL."""
    temp_file = download_pdf(url)
    if not temp_file:
        return []


    try:
        citations = extract_us_code_citations(temp_file, url)
    finally:
        os.remove(temp_file)  # Ensure the temporary file is deleted
    return citations


def save_to_excel(data, filename="extracted_citations.xlsx"):
    """Saves the extracted data to an Excel file."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Citation", "Citation Page", "Inferred Section Name", "Context", "URL"])


    for row in data:
        sanitized_row = [sanitize_text(str(cell)) for cell in row]
        sanitized_row[0] = clean_citation(sanitized_row[0])
        sheet.append(sanitized_row)


    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 20


    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2), start=2):
        for cell in row:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5), start=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


    workbook.save(filename)
    print(f"Saved data to {filename}")


def main():
    url_list = [
        "https://www.dhs.gov/sites/default/files/2024-09/2024_0923_cio_dhs_compliance_plan_omb_memoranda.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_047-01-privacy-policy-and-compliance_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/05.%20Directive%20138-01%2C%20Enterprise%20Information%20Technology%20Configuration%20Management%20%285-6-14%29.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_138-03-info-tech-asset-mgmt-and-refresh_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_139-02-info-quality_revision-01.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_139-02-001-info-quality-implementation_revision-01.pdf",
        "https://www.dhs.gov/sites/default/files/publications/139-05.pdf",
        "https://www.dhs.gov/sites/default/files/2023-09/23_0913_mgmt_139-06-acquistion-use-ai-technologies-dhs-components.pdf",
        "https://www.dhs.gov/sites/default/files/2023-11/23_1114_cio_use_generative_ai_tools.pdf",
        "https://www.dhs.gov/sites/default/files/2025-01/25_0116_CIO_DHS-Directive-139-08-508.pdf",
        "https://www.dhs.gov/sites/default/files/publications/Directive%20140-01%2C%20Revision%2002%2C%20Information%20Technology%20Security%20Program%20%28....pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_140-02-cybersecurity-workforce-mgmt-support_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_140-04-special-access-programs_revision-02.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_140-05-privacy-technology-implementation-guide.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_140-06-privacy-policy-research-programs-projects_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_141-01-records-and-information-management_revision-01.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_141-02-forms-management_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_142-01-information-collection-mgmt-program_revision-01.pdf",
        "https://www.dhs.gov/sites/default/files/publications/11.%20Directive%20142-02%20Information%20Technology%20Integration%20and%20Management%20%282-6-2014%29.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_142-03-electronic-mail-usage-and-maintenance_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-01-comp-match-agreements-data-integrity-board_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-02-disclose-asylum-refugee-info-c-terror-intel-purpose_rev-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-03-dhs-info-sharing-environment-tech-program_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-04-dhs-web-internet-extranet-information_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/2023-08/mgmt-dir_262-04-001-dhs-web-internet-extranet-information_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-05-information-sharing-and-safeguarding.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-06-digital-government-strategy_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-07-disclosure-of-homeland-security-info_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-08-protected-critical-infrastructure-info-program_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-09-enterprise-info-tech-service-management_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/DHS%20Digital%20Transformation.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-11-freedom-of-information-act-compliance_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-12-lexicon-program-standardization-dept-terminology_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-13-dhs-data-framework-terms-and-conditions.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-14-roles-and-responsibilities-for-shared-it-services.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_262-15-dhs-fed-info-share-enviro-privacy-civ-lib-protection-pol.pdf",
        "https://www.dhs.gov/sites/default/files/2022-05/mgmt-dir_262-16-00-privacy-policy-regarding-collection-use-retention-dissemination-pii.pdf",
        "https://www.dhs.gov/sites/default/files/2023-08/23_0810_mgmt_social-media-thrird-party-services-262-19.pdf",
        "https://www.dhs.gov/sites/default/files/2023-08/23_0803_mgmt_social-media-thrird-party-services-262-19-001.pdf",
        "https://www.dhs.gov/sites/default/files/2024-09/2024_0923_cio_dhs_compliance_plan_omb_memoranda.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_md-4100-1-wireless-management-office.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_md-4600-1-personal-use-of-government-office-equipment.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_md-4700-1-personal-communications-device-distribution.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_md-4800-telecommunications-operations.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_md-4900-individual-use-operation-dhs-info-systems-computers.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_047-01-privacy-policy-and-compliance_revision-00.pdf"
    ]


    all_citations = []
    for url in url_list:
        all_citations.extend(process_url(url))


    save_to_excel(all_citations)


if __name__ == "__main__":
    main()
