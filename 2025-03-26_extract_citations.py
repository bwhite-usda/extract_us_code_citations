# Filename: extract_us_code_citations_2025-03-26.py


import requests
import PyPDF2
import re
import os
import tempfile
import time
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def sanitize_text(text):
    return re.sub(r"[\r\n]+", " ", text).strip()


def clean_citation(citation):
    citation = re.sub(r"\b(\d+)\s*(U\.S\.C\.|USC)\s*(\d+)\b", r"\1 USC \3", citation)
    citation = re.sub(r"\b(\d+)\s*(C\.F\.R\.|CFR)\s*(\d+)\b", r"\1 CFR \3", citation)
    citation = re.sub(r"\b(E\.O\.|Executive\s*Order)\s*(\d+)\b", r"Executive Order \2", citation)
    citation = re.sub(r"\bEO\s+(\d+)\b", r"EO \1", citation)
    return citation


def get_browser_headers():
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "application/pdf",
        "Connection": "keep-alive"
    }


def download_pdf(url):
    try:
        session = requests.Session()
        retries = Retry(
            total=5,
            backoff_factor=5,
            status_forcelist=[500, 502, 503, 504],
            raise_on_status=False,
        )
        session.mount('https://', HTTPAdapter(max_retries=retries))


        response = session.get(url, headers=get_browser_headers(), stream=True, timeout=60)
        response.raise_for_status()


        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        for chunk in response.iter_content(chunk_size=1024):
            temp_file.write(chunk)
        temp_file.close()


        print(f"Downloaded {url}")
        return temp_file.name
    except Exception as e:
        print(f"Failed to download {url}: {e}")
        with open("failed_downloads.txt", "a") as f:
            f.write(url + "\n")
        return None


def extract_toc(reader):
    toc = []
    toc_pattern = r"(?P<heading>.+?)\s+(\d+)"
    for page_num, page in enumerate(reader.pages[:10]):
        text = page.extract_text()
        if text and "Table of Contents" in text:
            matches = re.findall(toc_pattern, text)
            for match in matches:
                heading = sanitize_text(match[0])
                page_start = int(match[1])
                toc.append((heading, page_start))
    return toc


def infer_section_name(toc, page_num, context, page_text):
    if toc:
        for i, (section, start_page) in enumerate(toc):
            if i + 1 < len(toc) and toc[i + 1][1] > page_num >= start_page:
                return section
            elif i == len(toc) - 1 and page_num >= start_page:
                return section
    lines = page_text.splitlines()
    context_start = page_text.find(context)
    for i in range(len(lines) - 1, -1, -1):
        if len(lines[i].strip()) > 0 and lines[i].strip() in page_text[:context_start]:
            return sanitize_text(lines[i])
    return "Unknown Section"


def extract_us_code_citations(pdf_path, url):
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            toc = extract_toc(reader)
            num_pages = len(reader.pages)
            citations = []


            citation_pattern = (
                r"\b(\d+)\s*(U\.S\.C\.|USC|U\.S\. Code)\s*\u00a7?\s*(\d+(\.\d+)*([a-zA-Z0-9]*)?)|"
                r"\b(\d+)\s*(C\.F\.R\.|CFR|Code of Federal Regulations)\s*\u00a7?\s*(\d+(\.\d+)*([a-zA-Z0-9]*)?)|"
                r"(E\.O\.|Executive\s*Order)\s*(\d+)|"
                r"\bEO\s+(\d+)\b"
            )


            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text = page.extract_text()
                if text:
                    matches = re.finditer(citation_pattern, text, re.IGNORECASE)
                    for match in matches:
                        citation_text = match.group(0)
                        citation_number = match.group(10) or match.group(8)
                        if citation_number:
                            citation_text = f"EO {citation_number}"
                        citation = clean_citation(citation_text)
                        start, end = match.start(), match.end()
                        context = sanitize_text(text[max(0, start - 100):min(len(text), end + 100)])
                        section_name = infer_section_name(toc, page_num + 1, context, text)
                        citation_page_url = f"{url}#page={page_num + 1}"
                        citations.append((citation, citation_page_url, section_name, context, url))
        return citations
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
        return []


def process_url(url):
    temp_file = download_pdf(url)
    if not temp_file:
        return []
    try:
        return extract_us_code_citations(temp_file, url)
    finally:
        os.remove(temp_file)


def save_to_excel(data, filename="extracted_citations.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Citation", "Citation Page", "Inferred Section Name", "Context", "URL"])


    for row in data:
        sanitized_row = [sanitize_text(str(cell)) for cell in row]
        sanitized_row[0] = clean_citation(sanitized_row[0])
        sheet.append(sanitized_row)


    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


    workbook.save(filename)
    print(f"Saved data to {filename}")


def main():
    url_list = [
        "https://www.usda.gov/sites/default/files/documents/DM3020-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3050-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3050-002.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3060-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3060-002.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR 3080-001 Records Management.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3085-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3090-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/REMOVAL OF RECORDS BY EMPLOYEES AND POLITICAL APPOINTEES.pdf",
        "https://www.usda.gov/sites/default/files/documents/dr-3105-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DM 3107-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR-3107-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3111-001_USDA IT Strategic Plan Process_FINAL.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR 3130-008_Definition of Major Information Technology Investments_Final.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3130-009_Non Major Information Technology Investments_FINAL.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3130-010_USDA Enterprise Information Technology Governance (EITG)_FINAL.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR 3130-011 IT Project and Program Managers Certification Requirements final.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3130-012.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3130-013.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3145-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3150-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DM3160-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3160-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3170-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DM3180-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3180-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3185-001.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3185-002.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3185-003.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3185-004.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-001-A.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-001-B.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-001-C.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-001-E.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-001-G.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-001-I.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-001-J.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-001-K.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-001-M.pdf",
        "https://www.usda.gov/sites/default/files/documents/DR3300-004.pdf"
    ]


    all_citations = []
    for url in url_list:
        all_citations.extend(process_url(url))
        time.sleep(3)  # pause between downloads to mimic human browsing


    save_to_excel(all_citations)


if __name__ == "__main__":
    main()
