

# This is extract_us_code_citations_2025-03-14.py


import requests
import PyPDF2
import re
import os
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def download_pdf(url):
    """Downloads a PDF from the given URL and saves it as a temporary file."""
    try:
        response = requests.get(url, stream=True, timeout=10)
        response.raise_for_status()
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        for chunk in response.iter_content(chunk_size=1024):
            temp_file.write(chunk)
        temp_file.close()
        print(f"Downloaded {url}")
        return temp_file.name
    except Exception as e:
        print(f"Failed to download {url}: {e}")
        return None


def sanitize_text(text):
    """Removes line breaks and carriage returns from a given text."""
    return re.sub(r"[\r\n]+", " ", text).strip()


def clean_citation(citation):
    """Cleans up citation syntax to match the format ## USC ## and Executive Order formatting."""
    citation = re.sub(r"\b(\d+)\s*(U\.S\.C\.|USC)\s*(\d+)\b", r"\1 USC \3", citation)
    citation = re.sub(r"\b(E\.O\.|Executive Order)\s*(\d+)\b", r"Executive Order \2", citation)
    return citation


def extract_toc(reader):
    """Attempts to extract the Table of Contents (TOC) from the PDF."""
    toc = []
    toc_pattern = r"(?P<heading>.+?)\s+(\d+)"
    for page_num, page in enumerate(reader.pages[:10]):  # Check the first 10 pages for a TOC
        text = page.extract_text()
        if text and "Table of Contents" in text:
            matches = re.findall(toc_pattern, text)
            for match in matches:
                heading = sanitize_text(match[0])
                page_start = int(match[1])
                toc.append((heading, page_start))
    return toc


def infer_section_name(toc, page_num, context, page_text):
    """Infers the section name based on TOC or contextual headers."""
    if toc:
        for i, (section, start_page) in enumerate(toc):
            if i + 1 < len(toc) and toc[i + 1][1] > page_num >= start_page:
                return section
            elif i == len(toc) - 1 and page_num >= start_page:  # Last TOC entry
                return section


    # Fallback: Find the nearest one-line paragraph before the context
    lines = page_text.splitlines()
    context_start = page_text.find(context)
    for i in range(len(lines) - 1, -1, -1):
        if len(lines[i].strip()) > 0 and lines[i].strip() in page_text[:context_start]:
            return sanitize_text(lines[i])
    return "Unknown Section"


def extract_us_code_citations(pdf_path, url):
    """Extracts U.S. Code, CFR, and Executive Order citations from a given PDF file."""
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            toc = extract_toc(reader)
            num_pages = len(reader.pages)


            citations = []
            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text = page.extract_text()


                if text:
                    citation_pattern = r"\b(\d+)\s*(U\.S\.C\.|USC|CFR|Code of Federal Regulations)\s*\u00a7?\s*(\d+(\.\d+)*([a-zA-Z0-9]*)?)|" \
                                       r"\b(E\.O\.|Executive Order)\s*(\d+)\b"


                    matches = re.finditer(citation_pattern, text)


                    for match in matches:
                        citation = clean_citation(match.group(0))
                        start, end = match.start(), match.end()
                        context = sanitize_text(text[max(0, start - 100):min(len(text), end + 100)])
                        section_name = infer_section_name(toc, page_num + 1, context, text)
                        citation_page_url = f"{url}#page={page_num + 1}"
                        citations.append((citation, citation_page_url, section_name, context, url))


            return citations
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
        return []


def process_url(url):
    """Handles downloading, extracting, and cleaning up for a single URL."""
    temp_file = download_pdf(url)
    if not temp_file:
        return []


    try:
        citations = extract_us_code_citations(temp_file, url)
    finally:
        os.remove(temp_file)  # Ensure the temporary file is deleted
    return citations


def save_to_excel(data, filename="extracted_citations.xlsx"):
    """Saves the extracted data to an Excel file."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Citation", "Citation Page", "Inferred Section Name", "Context", "URL"])


    for row in data:
        sanitized_row = [sanitize_text(str(cell)) for cell in row]
        sanitized_row[0] = clean_citation(sanitized_row[0])
        sheet.append(sanitized_row)


    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 20


    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2), start=2):
        for cell in row:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5), start=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


    workbook.save(filename)
    print(f"Saved data to {filename}")


def main():
    url_list = [
        "https://www.whitehouse.gov/wp-content/uploads/2018/06/a11.pdf",
        "https://www.whitehouse.gov/wp-content/uploads/legacy_drupal_files/omb/circulars/A123/a123_rev.pdf"
    ]


    all_citations = []
    for url in url_list:
        all_citations.extend(process_url(url))


    save_to_excel(all_citations)


if __name__ == "__main__":
    main()






